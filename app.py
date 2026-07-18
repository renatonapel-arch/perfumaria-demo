"""Perfumaria — backend Flask + SQLite

Endpoints REST usados pelo mockup HTML servido em `/`.
Persistência em `/data/perfumaria.db` (volume Coolify).
Scrape Fragrantica usa cache local por enquanto — troca por Playwright na v2.
"""
import os, json, sqlite3, uuid, datetime as dt
from pathlib import Path
from flask import Flask, jsonify, request, send_from_directory, abort

DB_PATH = Path(os.environ.get("DB_PATH", "/data/perfumaria.db"))
STATIC_DIR = Path(__file__).parent / "static"

app = Flask(__name__, static_folder=None)

# ---------------------------------------------------------------- db
def db():
    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row
    con.execute("PRAGMA foreign_keys=ON")
    return con

def init_db():
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    con = db()
    con.executescript("""
    CREATE TABLE IF NOT EXISTS perfume (
      id INTEGER PRIMARY KEY AUTOINCREMENT,
      marca TEXT NOT NULL,
      modelo TEXT NOT NULL,
      ano INTEGER,
      concentracao TEXT DEFAULT 'EDT',
      tamanho_ml INTEGER DEFAULT 100,
      estoque_ml REAL DEFAULT 100,
      ml_por_spray REAL DEFAULT 0.1,
      preco_pago REAL,
      familia TEXT DEFAULT 'oriental',
      notas_topo TEXT,
      notas_coracao TEXT,
      notas_fundo TEXT,
      rating INTEGER DEFAULT 0,
      nota_fragrantica REAL,
      votos_fragrantica INTEGER,
      review TEXT,
      referencia_id INTEGER,
      eh_wishlist INTEGER DEFAULT 0,
      ja_tive INTEGER DEFAULT 1,
      ja_terminou INTEGER DEFAULT 0,
      foto TEXT,
      criado_em TEXT DEFAULT CURRENT_TIMESTAMP,
      FOREIGN KEY(referencia_id) REFERENCES perfume(id) ON DELETE SET NULL
    );
    CREATE TABLE IF NOT EXISTS usage_log (
      id INTEGER PRIMARY KEY AUTOINCREMENT,
      perfume_id INTEGER NOT NULL,
      data TEXT NOT NULL,
      ocasiao TEXT,
      estacao TEXT,
      hora_dia TEXT,
      ml REAL DEFAULT 0.1,
      FOREIGN KEY(perfume_id) REFERENCES perfume(id) ON DELETE CASCADE
    );
    CREATE INDEX IF NOT EXISTS idx_usage_perfume ON usage_log(perfume_id);
    """)
    con.commit()
    if con.execute("SELECT COUNT(*) FROM perfume").fetchone()[0] == 0:
        con.execute("PRAGMA foreign_keys=OFF")
        _seed(con)
        con.execute("PRAGMA foreign_keys=ON")
    con.close()

def _seed(con):
    """Seed inicial com 10 perfumes + 18 sprays (mesmo do mockup)."""
    now = dt.datetime.now()
    def d(days, h, mi):
        t = now - dt.timedelta(days=days)
        return t.replace(hour=h, minute=mi, second=0, microsecond=0).isoformat()

    perfumes = [
        ('Zara', 'Vibrant Leather', 2018, 'EDT', 100, 62, 0.1, 149.90, 'leather',
         'Bergamota, toranja', 'Ameixa, rosa', 'Couro, almíscar, baunilha',
         4, 4.28, 3421, 'Impressionante o quanto se aproxima do Aventus. Longevidade média, projeção boa nas primeiras 4h.',
         4, 0, 1, 0, 'bottle'),
        ('Zara', 'Vibrant Leather Oud', 2020, 'EDP', 100, 78, 0.1, 189.90, 'oriental',
         'Cardamomo, bergamota', 'Oud, rosa, açafrão', 'Couro, âmbar, madeiras',
         5, 4.35, 1892, 'Melhor dupe Zara na minha coleção. Oud rico sem ficar sintético.',
         9, 0, 1, 0, 'bottle'),
        ('Zara', 'W/END 03:00 AM', 2022, 'EDP', 80, 45, 0.1, 129.90, 'oriental',
         'Pimenta rosa, cardamomo', 'Especiarias, íris', 'Baunilha, âmbar, madeiras',
         4, 4.05, 987, 'Vibe noturna, sedutor. Referência Black XS é evidente.',
         10, 0, 1, 0, 'bottle'),
        ('Creed', 'Aventus', 2010, 'EDP', 100, 88, 0.1, 3200.00, 'fresh',
         'Abacaxi, bergamota, maçã, cassis', 'Bétula, rosa, jasmim, patchouli',
         'Almíscar, musgo de carvalho, baunilha, âmbar cinza',
         5, 4.42, 28450, 'A referência. Padrão-ouro dos frutados masculinos. Usar com parcimônia.',
         None, 0, 1, 0, 'https://fimgs.net/mdimg/perfume/375x500.9828.jpg'),
        ('Chanel', 'Bleu de Chanel Parfum', 2018, 'Parfum', 100, 0, 0.1, 0, 'woody',
         'Bergamota, grapefruit, menta', 'Gengibre, jasmim, noz-moscada', 'Incenso, cedro, sândalo',
         5, 4.31, 45210, '',
         None, 1, 0, 0, 'bottle'),
        ('Prada', "L'Homme", 2016, 'EDT', 100, 0, 0.1, 420.00, 'floral',
         'Cardamomo, nerólio, pimenta preta', 'Íris, violeta, gerânio', 'Cedro, almíscar, âmbar cinza',
         4, 3.98, 5820, 'Elegante, sofisticado. Usava em jantares.',
         None, 0, 1, 1, 'bottle'),
        ('Dior', 'Fahrenheit', 1988, 'EDT', 100, 55, 0.1, 550.00, 'leather',
         'Espinheiro, mandarina, bergamota', 'Madressilva, sândalo, violeta', 'Couro, cedro, tonka',
         5, 4.15, 22103, 'Clássico atemporal. Nada como usar Fahrenheit no inverno.',
         None, 0, 1, 0, None),
        ('Yves Saint Laurent', "La Nuit de L'Homme", 2009, 'EDT', 100, 12, 0.1, 460.00, 'oriental',
         'Cardamomo', 'Lavanda, cominho, bergamota', 'Vetiver, cedro, cumarina',
         5, 4.29, 31890, 'Mais votos que qualquer outro. Especialmente forte no outono.',
         None, 0, 1, 0, 'bottle'),
        ('Tom Ford', 'Oud Wood', 2007, 'EDP', 50, 44, 0.1, 1890.00, 'oriental',
         'Pau-rosa, cardamomo, pimenta', 'Oud, sândalo, vetiver', 'Baunilha, âmbar, almíscar',
         5, 4.40, 12034, 'Elegância pura. Oud de alta qualidade.',
         None, 0, 1, 0, 'bottle'),
        ('Paco Rabanne', 'Black XS', 2005, 'EDT', 100, 0, 0.1, 320.00, 'oriental',
         'Toranja, álamo', 'Cardamomo, praliné', 'Patchouli, baunilha, tonka',
         4, 4.10, 15342, 'Meu perfume da adolescência. Fase 2000s pura.',
         None, 0, 1, 1, 'bottle'),
    ]
    cols = ("marca, modelo, ano, concentracao, tamanho_ml, estoque_ml, ml_por_spray, preco_pago, familia, "
            "notas_topo, notas_coracao, notas_fundo, rating, nota_fragrantica, votos_fragrantica, review, "
            "referencia_id, eh_wishlist, ja_tive, ja_terminou, foto")
    con.executemany(f"INSERT INTO perfume ({cols}) VALUES ({','.join(['?']*21)})", perfumes)

    sprays = [
        (1, d(1,8,15), 'trabalho', 'inverno', 'manhã', 0.1),
        (1, d(3,7,45), 'trabalho', 'inverno', 'manhã', 0.1),
        (1, d(5,19,20), 'social', 'inverno', 'noite', 0.2),
        (1, d(8,8,10), 'trabalho', 'inverno', 'manhã', 0.1),
        (1, d(12,20,30), 'noite', 'meia-estação', 'noite', 0.2),
        (1, d(18,7,50), 'trabalho', 'meia-estação', 'manhã', 0.1),
        (1, d(22,15,0), 'casa', 'meia-estação', 'tarde', 0.1),
        (1, d(28,19,15), 'social', 'verão', 'noite', 0.2),
        (7, d(2,20,0), 'noite', 'inverno', 'noite', 0.2),
        (7, d(6,19,45), 'social', 'inverno', 'noite', 0.1),
        (7, d(14,20,20), 'noite', 'inverno', 'noite', 0.2),
        (7, d(20,19,0), 'noite', 'meia-estação', 'noite', 0.2),
        (7, d(35,18,30), 'social', 'meia-estação', 'noite', 0.2),
        (7, d(45,20,0), 'noite', 'verão', 'noite', 0.1),
        (9, d(4,21,0), 'noite', 'inverno', 'noite', 0.1),
        (9, d(15,19,30), 'social', 'inverno', 'noite', 0.1),
        (9, d(25,20,0), 'noite', 'meia-estação', 'noite', 0.1),
        (9, d(40,21,15), 'noite', 'meia-estação', 'noite', 0.1),
    ]
    con.executemany("INSERT INTO usage_log (perfume_id, data, ocasiao, estacao, hora_dia, ml) VALUES (?,?,?,?,?,?)", sprays)
    con.commit()

def _row_to_dict(r):
    d = dict(r)
    d['id'] = str(d['id'])
    if d.get('referencia_id'):
        d['referencia_id'] = str(d['referencia_id'])
    d['eh_wishlist'] = bool(d.get('eh_wishlist'))
    d['ja_tive'] = bool(d.get('ja_tive'))
    d['ja_terminou'] = bool(d.get('ja_terminou'))
    return d

# ---------------------------------------------------------------- static
@app.get("/")
def index():
    return send_from_directory(STATIC_DIR, "app.html")

@app.get("/preview")
def preview():
    return send_from_directory(STATIC_DIR, "index.html")

@app.get("/app.html")
def app_html():
    return send_from_directory(STATIC_DIR, "app.html")

@app.get("/health")
def health():
    return {"ok": True, "db": DB_PATH.exists()}

# ---------------------------------------------------------------- perfumes
@app.get("/api/perfumes")
def list_perfumes():
    con = db()
    rows = con.execute("SELECT * FROM perfume ORDER BY id").fetchall()
    con.close()
    return jsonify([_row_to_dict(r) for r in rows])

@app.post("/api/perfumes")
def create_perfume():
    p = request.get_json() or {}
    if not p.get('marca') or not p.get('modelo'):
        return {"error": "marca e modelo obrigatórios"}, 400
    con = db()
    cur = con.execute("""INSERT INTO perfume
        (marca, modelo, ano, concentracao, tamanho_ml, estoque_ml, ml_por_spray, preco_pago,
         familia, notas_topo, notas_coracao, notas_fundo, rating, nota_fragrantica, votos_fragrantica,
         review, referencia_id, eh_wishlist, ja_tive, ja_terminou, foto)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""", (
        p.get('marca'), p.get('modelo'), p.get('ano'), p.get('concentracao','EDT'),
        p.get('tamanho_ml',100), p.get('estoque_ml', p.get('tamanho_ml',100)),
        p.get('ml_por_spray',0.1), p.get('preco_pago'),
        p.get('familia','oriental'), p.get('notas_topo'), p.get('notas_coracao'), p.get('notas_fundo'),
        p.get('rating',0), p.get('nota_fragrantica'), p.get('votos_fragrantica'),
        p.get('review'), p.get('referencia_id'),
        1 if p.get('eh_wishlist') else 0,
        0 if p.get('eh_wishlist') else 1,
        1 if p.get('ja_terminou') else 0,
        p.get('foto','bottle')
    ))
    new_id = cur.lastrowid
    con.commit()
    row = con.execute("SELECT * FROM perfume WHERE id=?", (new_id,)).fetchone()
    con.close()
    return jsonify(_row_to_dict(row))

@app.put("/api/perfumes/<int:pid>")
def update_perfume(pid):
    p = request.get_json() or {}
    fields = ['marca','modelo','ano','concentracao','tamanho_ml','estoque_ml','ml_por_spray',
              'preco_pago','familia','notas_topo','notas_coracao','notas_fundo','rating',
              'nota_fragrantica','votos_fragrantica','review','referencia_id','foto']
    sets = []
    vals = []
    for f in fields:
        if f in p:
            sets.append(f"{f}=?")
            vals.append(p[f])
    if 'eh_wishlist' in p:
        sets.append("eh_wishlist=?"); vals.append(1 if p['eh_wishlist'] else 0)
    if 'ja_terminou' in p:
        sets.append("ja_terminou=?"); vals.append(1 if p['ja_terminou'] else 0)
    if not sets: return {"error":"nada pra atualizar"}, 400
    vals.append(pid)
    con = db()
    con.execute(f"UPDATE perfume SET {', '.join(sets)} WHERE id=?", vals)
    con.commit()
    row = con.execute("SELECT * FROM perfume WHERE id=?", (pid,)).fetchone()
    con.close()
    return jsonify(_row_to_dict(row)) if row else ({"error":"not found"}, 404)

@app.delete("/api/perfumes/<int:pid>")
def delete_perfume(pid):
    con = db()
    con.execute("DELETE FROM perfume WHERE id=?", (pid,))
    con.commit()
    con.close()
    return {"ok": True}

# ---------------------------------------------------------------- spray
@app.post("/api/perfumes/<int:pid>/spray")
def register_spray(pid):
    body = request.get_json() or {}
    con = db()
    row = con.execute("SELECT estoque_ml, ml_por_spray, tamanho_ml FROM perfume WHERE id=?", (pid,)).fetchone()
    if not row: con.close(); return {"error":"not found"}, 404
    ml = float(body.get('ml') or row['ml_por_spray'])
    if row['estoque_ml'] < ml:
        con.close(); return {"error":"estoque insuficiente"}, 400
    new_est = round(row['estoque_ml'] - ml, 2)
    con.execute("UPDATE perfume SET estoque_ml=? WHERE id=?", (new_est, pid))
    con.execute("INSERT INTO usage_log (perfume_id, data, ocasiao, estacao, hora_dia, ml) VALUES (?,?,?,?,?,?)",
                (pid, dt.datetime.now().isoformat(), body.get('ocasiao','trabalho'),
                 body.get('estacao','meia-estação'), body.get('hora_dia','manhã'), ml))
    con.commit()
    perfume = con.execute("SELECT * FROM perfume WHERE id=?", (pid,)).fetchone()
    con.close()
    pct = round((new_est / row['tamanho_ml']) * 100)
    return jsonify({"perfume": _row_to_dict(perfume), "pct_restante": pct})

@app.get("/api/perfumes/<int:pid>/sprays")
def list_sprays(pid):
    con = db()
    rows = con.execute("SELECT * FROM usage_log WHERE perfume_id=? ORDER BY data DESC", (pid,)).fetchall()
    con.close()
    return jsonify([{**dict(r), 'id': str(r['id']), 'perfume_id': str(r['perfume_id'])} for r in rows])

@app.delete("/api/sprays/<int:sid>")
def delete_spray(sid):
    con = db()
    row = con.execute("SELECT perfume_id, ml FROM usage_log WHERE id=?", (sid,)).fetchone()
    if not row: con.close(); return {"error":"not found"}, 404
    perf = con.execute("SELECT estoque_ml, tamanho_ml FROM perfume WHERE id=?", (row['perfume_id'],)).fetchone()
    new_est = min(perf['tamanho_ml'], round(perf['estoque_ml'] + row['ml'], 2))
    con.execute("UPDATE perfume SET estoque_ml=? WHERE id=?", (new_est, row['perfume_id']))
    con.execute("DELETE FROM usage_log WHERE id=?", (sid,))
    con.commit()
    con.close()
    return {"ok": True, "estoque_ml": new_est}

# ---------------------------------------------------------------- scrape
FRAGRANTICA_CACHE = {
  'aventus': dict(marca='Creed', modelo='Aventus', ano=2010, concentracao='EDP', familia='fresh',
    notas_topo='Abacaxi, bergamota, maçã, cassis',
    notas_coracao='Bétula, rosa, jasmim, patchouli',
    notas_fundo='Almíscar, musgo de carvalho, baunilha, âmbar cinza',
    nota_fragrantica=4.42, votos_fragrantica=28450,
    foto='https://fimgs.net/mdimg/perfume/375x500.9828.jpg'),
  'sauvage-elixir': dict(marca='Dior', modelo='Sauvage Elixir', ano=2021, concentracao='Elixir', familia='oriental',
    notas_topo='Canela, noz-moscada, cardamomo, toranja',
    notas_coracao='Lavanda, alcaçuz',
    notas_fundo='Sândalo, âmbar cinza, patchouli, madeiras de haiti',
    nota_fragrantica=4.35, votos_fragrantica=18732,
    foto='https://fimgs.net/mdimg/perfume/375x500.63792.jpg'),
  'y-edp': dict(marca='Yves Saint Laurent', modelo='Y Eau de Parfum', ano=2018, concentracao='EDP', familia='woody',
    notas_topo='Maçã, gengibre, bergamota',
    notas_coracao='Sálvia, gerânio, lavanda',
    notas_fundo='Fava tonka, cedro, âmbar cinza',
    nota_fragrantica=4.19, votos_fragrantica=12580,
    foto='https://fimgs.net/mdimg/perfume/375x500.49232.jpg'),
  '1m-elixir': dict(marca='Paco Rabanne', modelo='1 Million Elixir', ano=2022, concentracao='Parfum', familia='gourmand',
    notas_topo='Grapefruit, canela, pimenta rosa',
    notas_coracao='Cardamomo, âmbar liquido',
    notas_fundo='Baunilha, fava tonka, patchouli',
    nota_fragrantica=4.02, votos_fragrantica=4260,
    foto='https://fimgs.net/mdimg/perfume/375x500.77441.jpg'),
  'adg-profondo': dict(marca='Giorgio Armani', modelo='Acqua di Giò Profondo', ano=2020, concentracao='EDP', familia='aquatic',
    notas_topo='Bergamota, notas aquáticas, notas marinhas, limão',
    notas_coracao='Rosmarinho, lavanda, alecrim, cipreste',
    notas_fundo='Patchouli, âmbar cinza, almíscar, musgo',
    nota_fragrantica=4.31, votos_fragrantica=15920,
    foto='https://fimgs.net/mdimg/perfume/375x500.59692.jpg'),
  'tobacco-vanille': dict(marca='Tom Ford', modelo='Tobacco Vanille', ano=2007, concentracao='EDP', familia='tobacco',
    notas_topo='Folha de tabaco, especiarias',
    notas_coracao='Baunilha, cacau, fruta de tabaco, fava tonka',
    notas_fundo='Madeiras secas frutadas',
    nota_fragrantica=4.38, votos_fragrantica=21340,
    foto='https://fimgs.net/mdimg/perfume/375x500.1825.jpg'),
}

def _match_scrape_key(url):
    u = (url or '').lower()
    if 'creed' in u and 'aventus' in u: return 'aventus'
    if 'sauvage-elixir' in u: return 'sauvage-elixir'
    if 'y-eau-de-parfum' in u or ('yves-saint-laurent' in u and '/y-' in u): return 'y-edp'
    if '1-million-elixir' in u: return '1m-elixir'
    if 'acqua-di-gio-profondo' in u: return 'adg-profondo'
    if 'tobacco-vanille' in u: return 'tobacco-vanille'
    return None

@app.post("/api/scrape")
def scrape():
    body = request.get_json() or {}
    url = (body.get('url') or '').strip()
    if not url: return {"error":"URL vazia"}, 400
    if 'fragrantica' not in url.lower(): return {"error":"URL precisa ser do Fragrantica"}, 400
    # 1) tenta scraper real (Playwright)
    try:
        from scraper import scrape_fragrantica, ScrapeError
        data = scrape_fragrantica(url)
        return jsonify(data)
    except ScrapeError as e:
        # 2) fallback pro cache (6 perfumes conhecidos)
        key = _match_scrape_key(url)
        cached = FRAGRANTICA_CACHE.get(key)
        if cached: return jsonify(cached)
        return {"error": f"scrape falhou e sem cache: {e}"}, 502
    except Exception as e:
        # 3) fallback silencioso pro cache mesmo em erro inesperado
        key = _match_scrape_key(url)
        cached = FRAGRANTICA_CACHE.get(key)
        if cached: return jsonify(cached)
        return {"error": f"erro inesperado: {e}"}, 500

# ---------------------------------------------------------------- random
@app.get("/api/random-pick")
def random_pick():
    import random
    ocasiao = request.args.get('ocasiao','qualquer')
    con = db()
    rows = con.execute(
        "SELECT * FROM perfume WHERE ja_tive=1 AND ja_terminou=0 AND eh_wishlist=0 AND estoque_ml >= ml_por_spray"
    ).fetchall()
    con.close()
    if not rows: return {"error":"nenhum disponível"}, 404
    return jsonify(_row_to_dict(random.choice(rows)))

# ---------------------------------------------------------------- import xlsx real
import unicodedata
def _norm(s):
    if s is None: return ''
    s = str(s).strip().lower()
    return ''.join(c for c in unicodedata.normalize('NFKD', s) if not unicodedata.combining(c))

# possíveis nomes de coluna → campo interno
COL_MAP = {
    'marca': ['marca', 'brand', 'fabricante', 'casa'],
    'modelo': ['modelo', 'nome', 'name', 'perfume', 'fragrancia', 'fragrance', 'produto'],
    'ano': ['ano', 'year', 'lancamento', 'lançamento'],
    'concentracao': ['concentracao', 'concentração', 'tipo', 'edt', 'edp', 'type'],
    'tamanho_ml': ['tamanho', 'ml', 'tamanho_ml', 'tamanho (ml)', 'volume', 'size'],
    'estoque_ml': ['estoque', 'estoque_ml', 'restante', 'saldo'],
    'ml_por_spray': ['ml_por_spray', 'ml/spray', 'ml por spray', 'consumo'],
    'preco_pago': ['preco', 'preço', 'preco_pago', 'valor', 'custo', 'preço pago', 'price'],
    'familia': ['familia', 'família', 'family', 'olfativa', 'família olfativa'],
    'notas_topo': ['notas_topo', 'topo', 'top', 'notas de topo', 'top notes'],
    'notas_coracao': ['notas_coracao', 'coração', 'coracao', 'heart', 'middle', 'notas de coração'],
    'notas_fundo': ['notas_fundo', 'fundo', 'base', 'base notes', 'notas de fundo'],
    'rating': ['rating', 'nota_pessoal', 'nota pessoal', 'minha_nota', 'stars', 'estrelas'],
    'nota_fragrantica': ['nota_fragrantica', 'fragrantica', 'nota fragrantica', 'nota fragantica'],
    'votos_fragrantica': ['votos', 'votos_fragrantica', 'reviews'],
    'review': ['review', 'anotacao', 'anotação', 'nota', 'observacao', 'observação', 'comentario', 'comentário'],
    'eh_wishlist': ['wishlist', 'quero', 'lista de desejos', 'desejo'],
    'ja_terminou': ['terminou', 'ja_terminou', 'ja tive', 'já tive', 'antigo', 'arquivo', 'finalizado'],
    'foto': ['foto', 'imagem', 'photo', 'image', 'url'],
}

def _detect_columns(headers):
    """Devolve {campo_interno: idx_coluna} baseado em fuzzy match de cabeçalhos."""
    out = {}
    for idx, h in enumerate(headers):
        nh = _norm(h)
        if not nh: continue
        for field, aliases in COL_MAP.items():
            if field in out: continue  # 1º match vence
            if any(_norm(a) == nh or _norm(a) in nh for a in aliases):
                out[field] = idx
                break
    return out

def _cell_bool(v):
    if v is None: return None
    s = _norm(v)
    if s in ('sim', 's', 'yes', 'y', 'true', '1', 'x'): return True
    if s in ('nao', 'não', 'n', 'no', 'false', '0', ''): return False
    return None

def _cell_num(v):
    if v is None: return None
    if isinstance(v, (int, float)): return v
    s = str(v).strip().replace('R$','').replace(' ','').replace(',','.')
    try: return float(s)
    except (ValueError, TypeError): return None

@app.post("/api/import")
def import_xlsx():
    from openpyxl import load_workbook
    from io import BytesIO
    mode = request.form.get('mode', 'add')
    if 'file' not in request.files:
        return {"error": "arquivo não enviado (campo 'file')"}, 400
    f = request.files['file']
    try:
        wb = load_workbook(BytesIO(f.read()), data_only=True, read_only=True)
    except Exception as e:
        return {"error": f"não consegui abrir a planilha: {e}"}, 400
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return {"error": "planilha vazia ou só com cabeçalho"}, 400
    # descobre header row: primeira row com marca+modelo (ou similares)
    header_row_idx = 0
    header_cols = {}
    for i in range(min(5, len(rows))):
        candidate = _detect_columns(rows[i])
        if 'marca' in candidate and 'modelo' in candidate:
            header_row_idx = i
            header_cols = candidate
            break
    if not header_cols:
        return {"error": "não achei colunas 'marca' e 'modelo' — cheque o cabeçalho"}, 400

    con = db()
    if mode == 'replace':
        con.execute("DELETE FROM usage_log")
        con.execute("DELETE FROM perfume")

    inserted = 0
    skipped = 0
    for row in rows[header_row_idx + 1:]:
        if not row: continue
        def _v(field):
            i = header_cols.get(field)
            return row[i] if i is not None and i < len(row) else None
        marca = _v('marca')
        modelo = _v('modelo')
        if not marca or not modelo:
            skipped += 1
            continue
        tamanho = _cell_num(_v('tamanho_ml')) or 100
        estoque = _cell_num(_v('estoque_ml'))
        if estoque is None: estoque = tamanho
        eh_wish = _cell_bool(_v('eh_wishlist')) or False
        ja_term = _cell_bool(_v('ja_terminou')) or False
        con.execute("""INSERT INTO perfume
            (marca, modelo, ano, concentracao, tamanho_ml, estoque_ml, ml_por_spray, preco_pago,
             familia, notas_topo, notas_coracao, notas_fundo, rating, nota_fragrantica, votos_fragrantica,
             review, eh_wishlist, ja_tive, ja_terminou, foto)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""", (
            str(marca).strip(), str(modelo).strip(),
            int(_cell_num(_v('ano')) or 0) or None,
            str(_v('concentracao') or 'EDT').strip(),
            int(tamanho), float(estoque),
            float(_cell_num(_v('ml_por_spray')) or 0.1),
            _cell_num(_v('preco_pago')),
            str(_v('familia') or 'oriental').strip().lower()[:20],
            _v('notas_topo'), _v('notas_coracao'), _v('notas_fundo'),
            int(_cell_num(_v('rating')) or 0),
            _cell_num(_v('nota_fragrantica')),
            int(_cell_num(_v('votos_fragrantica')) or 0) or None,
            _v('review'),
            1 if eh_wish else 0,
            0 if eh_wish else 1,
            1 if ja_term else 0,
            _v('foto') or 'bottle',
        ))
        inserted += 1
    con.commit()
    con.close()
    return {"ok": True, "imported": inserted, "skipped": skipped, "columns_matched": list(header_cols.keys())}

# ---------------------------------------------------------------- boot
init_db()

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT",5006)), debug=True)
